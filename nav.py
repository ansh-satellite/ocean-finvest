"""
nav_updater.py
==============
Drop this file in your Ocean Finvest folder (same directory as Dash.py).

Handles:
  - Reading existing NAV.xlsx
  - Fetching BSE500 level via TrueData
  - Appending new rows (chained base-100 NAV) without overwriting history
  - Recomputing CalendarReturns sheet automatically
  - Streamlit UI section (call render_nav_section() from Dash.py)
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import streamlit as st

# ─────────────────────────────────────────────────────────────
# CONFIG  — edit path to match your machine
# ─────────────────────────────────────────────────────────────
NAV_FILE_PATH = r"C:\Users\LENOVO\Desktop\Ocean Finvest\NAV.xlsx"
TRUEDATA_USERNAME = "tdwsf695"
TRUEDATA_PASSWORD = "ocean@695"


# ═════════════════════════════════════════════════════════════
# 1.  FILE I/O
# ═════════════════════════════════════════════════════════════

def load_nav(path=NAV_FILE_PATH) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Sheet1")
    df.columns = df.columns.str.strip()
    df["DATE"] = pd.to_datetime(df["DATE"])
    return df.sort_values("DATE").reset_index(drop=True)


def _append_rows_to_file(path, new_rows: pd.DataFrame):
    """Append new NAV rows to Sheet1 without touching existing data."""
    wb = load_workbook(path)
    ws = wb["Sheet1"]
    last_row = ws.max_row

    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for _, row in new_rows.iterrows():
        last_row += 1
        c1 = ws.cell(row=last_row, column=1, value=row["DATE"].date())
        c2 = ws.cell(row=last_row, column=2, value=round(float(row["PORT NAV"]), 5))
        c3 = ws.cell(row=last_row, column=3, value=round(float(row["BM NAV"]),   5))
        c1.number_format = "DD-MMM-YYYY"
        c2.number_format = "0.00000"
        c3.number_format = "0.00000"
        for c in (c1, c2, c3):
            c.border = border

    wb.save(path)


def _write_calendar_sheet(path, full_df: pd.DataFrame):
    """Recreate the CalendarReturns sheet from full NAV history."""
    cal = _compute_calendar(full_df)

    wb = load_workbook(path)
    if "CalendarReturns" in wb.sheetnames:
        del wb["CalendarReturns"]
    ws = wb.create_sheet("CalendarReturns")

    hdr_fill = PatternFill("solid", start_color="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    pos_fill = PatternFill("solid", start_color="C6EFCE")
    neg_fill = PatternFill("solid", start_color="FFC7CE")
    thin     = Side(style="thin")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Month", "PORT Ret%", "BM Ret%", "Alpha%"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border    = border

    for ri, row in cal.iterrows():
        er = ri + 2
        for ci, col in enumerate(headers, 1):
            cell = ws.cell(row=er, column=ci, value=row[col])
            cell.border    = border
            cell.alignment = Alignment(horizontal="center")
            if ci == 1:
                cell.font = Font(bold=True, name="Arial")
            else:
                cell.number_format = '0.00"%"'
                if isinstance(row[col], (int, float)):
                    cell.fill = pos_fill if row[col] >= 0 else neg_fill

    ws.column_dimensions["A"].width = 14
    for col_letter in ["B", "C", "D"]:
        ws.column_dimensions[col_letter].width = 13

    wb.save(path)
    return cal


# ═════════════════════════════════════════════════════════════
# 2.  NAV CHAIN CALCULATION
# ═════════════════════════════════════════════════════════════

def _chain_nav(existing_df, nav_date, today_port, prev_port, today_bm, prev_bm):
    prev_port_nav = existing_df["PORT NAV"].iloc[-1]
    prev_bm_nav   = existing_df["BM NAV"].iloc[-1]
    port_ret = (today_port - prev_port) / prev_port if prev_port else 0
    bm_ret   = (today_bm   - prev_bm)   / prev_bm   if prev_bm   else 0
    return pd.DataFrame([{
        "DATE":     pd.Timestamp(nav_date),
        "PORT NAV": prev_port_nav * (1 + port_ret),
        "BM NAV":   prev_bm_nav   * (1 + bm_ret),
    }])


def _compute_calendar(df: pd.DataFrame) -> pd.DataFrame:
    """
    For each month:
      - Find the first available date in that month
      - Use the NAV of the day just before that date as the base
      - Use the last available date in that month as the end
      Return = (end NAV - base NAV) / base NAV * 100
    For the very first month (no prior data), first date of month is used as base.
    For current incomplete month, latest available row is used as end.
    """
    df = df.copy()
    df["YearMonth"] = df["DATE"].dt.to_period("M")
    rows = []
    for ym, grp in df.groupby("YearMonth"):
        grp = grp.sort_values("DATE")
        first_date = grp["DATE"].iloc[0]
        prev_rows  = df[df["DATE"] < first_date]
        if prev_rows.empty:
            sp, sb = grp["PORT NAV"].iloc[0], grp["BM NAV"].iloc[0]
        else:
            sp, sb = prev_rows["PORT NAV"].iloc[-1], prev_rows["BM NAV"].iloc[-1]
        ep, eb = grp["PORT NAV"].iloc[-1], grp["BM NAV"].iloc[-1]
        pr = round((ep - sp) / sp * 100, 2) if sp else 0
        br = round((eb - sb) / sb * 100, 2) if sb else 0
        rows.append({"Month": str(ym), "PORT Ret%": pr, "BM Ret%": br, "Alpha%": round(pr - br, 2)})
    return pd.DataFrame(rows)


# ═════════════════════════════════════════════════════════════
# 3.  TRUEDATA — FETCH BSE500 EOD LEVEL
# ═════════════════════════════════════════════════════════════

def fetch_bse500_level() -> tuple:
    """Returns (close_price: float | None, error: str | None)"""
    try:
        from truedata_ws.websocket.TD import TD
    except ImportError:
        return None, "truedata-ws not installed"

    symbols = ["BSE500", "S&P BSE 500", "CNX500", "NIFTY500"]
    td = None
    try:
        td = TD(TRUEDATA_USERNAME, TRUEDATA_PASSWORD, live_port=None)
        for sym in symbols:
            try:
                bars = td.get_n_historical_bars(sym, no_of_bars=1, bar_size="eod")
                if bars:
                    row = bars[-1]
                    for key in ["close", "Close", "c", "ltp", "LTP"]:
                        val = row.get(key) if isinstance(row, dict) else getattr(row, key, None)
                        if val is not None:
                            return float(val), None
            except Exception:
                continue
        return None, f"No data from symbols: {symbols}"
    except Exception as e:
        return None, str(e)
    finally:
        try:
            if td: td.disconnect()
        except Exception:
            pass


# ═════════════════════════════════════════════════════════════
# 4.  MAIN UPDATE FUNCTION
# ═════════════════════════════════════════════════════════════

def append_today_nav(today_port, prev_port, bse_today, bse_prev,
                     nav_date=None, path=NAV_FILE_PATH):
    """
    Append one new chained NAV row and refresh CalendarReturns sheet.
    Returns (success: bool, message: str, calendar_df or None)
    """
    if nav_date is None:
        nav_date = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    existing  = load_nav(path)
    last_date = existing["DATE"].max()

    if pd.Timestamp(nav_date) <= last_date:
        return False, (
            f"⚠️ {pd.Timestamp(nav_date).strftime('%d %b %Y')} already exists "
            f"(last entry: {last_date.strftime('%d %b %Y')}). Nothing appended."
        ), None

    new_row = _chain_nav(existing, nav_date, today_port, prev_port, bse_today, bse_prev)
    _append_rows_to_file(path, new_row)

    full_df = load_nav(path)
    cal_df  = _write_calendar_sheet(path, full_df)

    msg = (f"✅ Appended {pd.Timestamp(nav_date).strftime('%d %b %Y')}  |  "
           f"PORT NAV: {new_row['PORT NAV'].iloc[0]:.5f}  |  "
           f"BM NAV: {new_row['BM NAV'].iloc[0]:.5f}")
    return True, msg, cal_df


# ═════════════════════════════════════════════════════════════
# 5.  STREAMLIT UI  — call render_nav_section() from Dash.py
# ═════════════════════════════════════════════════════════════

def _colour(val):
    try:
        v = float(val)
        c   = "#00e676" if v >= 0 else "#ff5252"
        sgn = "+" if v >= 0 else ""
        return f'<span style="color:{c};font-weight:600;">{sgn}{v:.2f}%</span>'
    except Exception:
        return str(val)


def _render_calendar_table(cal_df: pd.DataFrame):
    display = cal_df.copy()
    display["PORT Ret%"] = display["PORT Ret%"].apply(_colour)
    display["BM Ret%"]   = display["BM Ret%"].apply(_colour)
    display["Alpha%"]    = display["Alpha%"].apply(_colour)
    st.write(display.to_html(escape=False, index=False), unsafe_allow_html=True)


def render_nav_section(path: str = NAV_FILE_PATH):
    """
    Paste these two lines anywhere in Dash.py to add the NAV section:

        from nav_updater import render_nav_section
        render_nav_section()
    """
    st.markdown("---")
    st.markdown("## 📒 NAV Tracker — Daily Update")
    st.markdown('<div class="glassy-container">', unsafe_allow_html=True)

    # ── load & show current state ──────────────────────────
    try:
        existing      = load_nav(path)
        last_date     = existing["DATE"].max()
        last_port_nav = existing["PORT NAV"].iloc[-1]
        last_bm_nav   = existing["BM NAV"].iloc[-1]
    except Exception as e:
        st.error(f"Could not read NAV.xlsx: {e}")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    c1, c2, c3 = st.columns(3)
    c1.metric("Last NAV Date",  last_date.strftime("%d %b %Y"))
    c2.metric("PORT NAV",       f"{last_port_nav:.3f}")
    c3.metric("BM NAV",         f"{last_bm_nav:.3f}")

    st.markdown("---")
    st.markdown("#### ➕ Add Today's NAV")

    # ── date + portfolio inputs ────────────────────────────
    col_d, col_p, col_pp = st.columns(3)
    with col_d:
        nav_date = st.date_input(
            "Date to add",
            value=datetime.today().date(),
            help="Select the trading date",
        )
    with col_p:
        today_port = st.number_input(
            "Today's Portfolio Total (₹)",
            min_value=0.0, step=1000.0, format="%.2f",
            help="Sum of Buy_Hold_Value for all active holdings today",
        )
    with col_pp:
        prev_port = st.number_input(
            "Previous Day's Portfolio Total (₹)",
            min_value=0.0, step=1000.0, format="%.2f",
            help="Same sum for the previous trading day",
        )

    # ── BSE500 inputs ──────────────────────────────────────
    st.markdown("#### 📡 BSE500 Benchmark Level")
    col_btn, col_bse, col_bsep = st.columns(3)

    with col_btn:
        if st.button("🔄 Auto-Fetch BSE500 (TrueData)", use_container_width=True):
            with st.spinner("Fetching from TrueData..."):
                val, err = fetch_bse500_level()
            if val:
                st.session_state["_nav_bse_fetched"] = val
                st.success(f"✅ BSE500: {val:,.2f}")
            else:
                st.error(f"❌ {err}")

    with col_bse:
        bse_today = st.number_input(
            "BSE500 Today",
            min_value=0.0, step=0.01, format="%.2f",
            value=float(st.session_state.get("_nav_bse_fetched", 0.0)),
            help="Auto-filled after fetching, or enter manually",
        )
    with col_bsep:
        bse_prev = st.number_input(
            "BSE500 Previous Day",
            min_value=0.0, step=0.01, format="%.2f",
            help="BSE500 closing level of the previous trading day",
        )

    # ── submit ─────────────────────────────────────────────
    st.markdown("")
    if st.button("💾  Append to NAV.xlsx & Refresh Calendar", use_container_width=True):
        if any(v == 0 for v in [today_port, prev_port, bse_today, bse_prev]):
            st.warning("⚠️ All four values must be non-zero before saving.")
        else:
            nav_dt = datetime.combine(nav_date, datetime.min.time())
            with st.spinner("Writing to NAV.xlsx..."):
                ok, msg, cal_df = append_today_nav(
                    today_port, prev_port, bse_today, bse_prev,
                    nav_date=nav_dt, path=path,
                )
            if ok:
                st.success(msg)
                st.session_state.pop("_nav_bse_fetched", None)
                st.markdown("#### 📅 Updated Calendar Returns")
                _render_calendar_table(cal_df)
            else:
                st.warning(msg)

    # ── always show latest calendar ────────────────────────
    st.markdown("#### 📅 Calendar Returns")
    try:
        _render_calendar_table(_compute_calendar(load_nav(path)))
    except Exception as e:
        st.warning(f"Calendar error: {e}")

    st.markdown("</div>", unsafe_allow_html=True)